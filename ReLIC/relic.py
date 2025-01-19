import os
from tqdm import tqdm
import torch
import numpy as np
import shutil
from torch.utils.data import DataLoader
from scipy.stats import pearsonr
from scipy.stats import spearmanr
from tensorboardX import SummaryWriter
from sklearn.metrics import accuracy_score
from argumentsparser import args

# from models.u_model import NIMA
# from models.e_model import NIMA
# from models.relic_model import NIMA
# from models.relic1_model import NIMA
from .models.relic2_model import NIMA
from .dataset import AVADataset
from .util import EDMLoss,AverageMeter,save_excel,make_csv,save_dataset
from RIE.dataloader.anydataset import AnyDataset
from . import option
import openpyxl

f = open('ReLIC/data/log_test.txt', 'w')

opt = args
opt.device = torch.device("cuda:{}".format(opt.gpu_id))

def adjust_learning_rate(params, optimizer, epoch):
    """Sets the learning rate to the initial LR
       decayed by 10 every 30 epochs"""
    lr = params.init_lr * (0.1 ** (epoch // 10))
    for param_group in optimizer.param_groups:
        param_group['lr'] = lr

def get_score(opt,y_pred):
    w = torch.from_numpy(np.linspace(1,10, 10))
    w = w.type(torch.FloatTensor)
    w = w.to(opt.device)

    w_batch = w.repeat(y_pred.size(0), 1)

    score = (y_pred * w_batch).sum(dim=1)
    score_np = score.data.cpu().numpy()
    return score, score_np


def create_data_part(opt, i, editeds=False):
    test_csv_path = os.path.join(opt.path_to_save_csv, 'test.csv')

    if editeds:
        test_ds = AVADataset(test_csv_path, editeds, if_train=False)
    else:
        test_ds = AVADataset(test_csv_path, opt.path_to_images+"/"+str(i-1), if_train=False)
    
    test_loader = DataLoader(test_ds, batch_size=opt.batch_size, num_workers=opt.num_workers, shuffle=False, pin_memory=True)

    return test_loader


def train(opt,model, loader, optimizer, criterion, writer=None, global_step=None, name=None):
    model.train()
    train_losses = AverageMeter()
    for idx, (x, y) in enumerate(tqdm(loader)):
        x = x.to(opt.device)
        y = y.to(opt.device)
        y_pred = model(x)
        loss = criterion(p_target=y, p_estimate=y_pred)
        optimizer.zero_grad()
        loss.backward()
        optimizer.step()
        train_losses.update(loss.item(), x.size(0))

        if writer is not None:
            writer.add_scalar(f"{name}/train_loss.avg", train_losses.avg, global_step=global_step + idx)
    return train_losses.avg


def validate(opt, model, loader, writer=None, global_step=None, name=None):
    model.eval()
    pred_score = []
    for idx, (x, y) in enumerate(tqdm(loader)):
        x = x.to(opt.device)
        y = y.type(torch.FloatTensor)
        y = y.to(opt.device)

        y_pred = model(x)
        pscore, pscore_np = get_score(opt,y_pred)

        pred_score += pscore_np.tolist()
    pred_score = np.array(pred_score)

    return pred_score


def start_train(opt):
    train_loader, val_loader, test_loader = create_data_part(opt)
    model = NIMA()
    optimizer = torch.optim.Adam(model.parameters(), lr=opt.init_lr)
    # optimizer = torch.optim.Adam(filter(lambda p: p.requires_grad, model.parameters()), lr=opt.init_lr)
    criterion = EDMLoss()
    model = model.to(opt.device)
    criterion.to(opt.device)

    writer = SummaryWriter(log_dir=os.path.join(opt.experiment_dir_name, 'logs'))

    for e in range(opt.num_epoch):
        adjust_learning_rate(opt, optimizer, e)
        train_loss = train(opt,model=model, loader=train_loader, optimizer=optimizer, criterion=criterion,
                           writer=writer, global_step=len(train_loader) * e,
                           name=f"{opt.experiment_dir_name}_by_batch")
        val_loss,vacc,vlcc,vsrcc = validate(opt,model=model, loader=val_loader, criterion=criterion,
                            writer=writer, global_step=len(val_loader) * e,
                            name=f"{opt.experiment_dir_name}_by_batch")
        test_loss, tacc, tlcc, tsrcc = validate(opt, model=model, loader=test_loader, criterion=criterion,
                                               writer=writer, global_step=len(val_loader) * e,
                                               name=f"{opt.experiment_dir_name}_by_batch")
        model_name = f"epoch_{e}_.pth"
        torch.save(model.state_dict(), os.path.join(opt.experiment_dir_name, model_name))

        f.write(
            'epoch:%d,v_lcc:%.4f,v_srcc:%.4f,v_acc:%.5f,tlcc:%.4f,t_srcc:%.4f,t_acc:%.5f,train:%.5f,val:%.5f,test:%.5f\r\n'
            % (e, vlcc[0], vsrcc[0], vacc, tlcc[0], tsrcc[0], tacc, train_loss, val_loss, test_loss))
        f.flush()

        writer.add_scalars("epoch_loss", {'train': train_loss, 'val': val_loss, 'test': test_loss},
                           global_step=e)

        writer.add_scalars("lcc_srcc", {'val_lcc': vlcc[0], 'val_srcc': vsrcc[0],
                                        'test_lcc': tlcc[0], 'test_srcc': tsrcc[0]},
                           global_step=e)

        writer.add_scalars("acc",{'val_acc': vacc, 'test_acc': tacc}, global_step=e)

    writer.close()
    f.close()


def aesthetic_prediction(opt, model, batch_num, editeds):

    dataloader = create_data_part(opt, batch_num)
    pred_score = validate(opt,model=model, loader=dataloader)
  
    return pred_score

def saliency_prediction(args, trainer, batch_num):

    dataset = AnyDataset(args, batch_num)
    dataloader = torch.utils.data.DataLoader(
        dataset,
        batch_size=1,  # Change this to >1 if you can
        shuffle=False,
        num_workers=args.num_workers,
        pin_memory=True,
        drop_last=False
    )

    final_sal_list = []
    for episode, data in enumerate(dataloader):
        trainer.setinput_hr(data)

        # Initialize a NumPy array to hold the results for this batch
        sal_batch = np.zeros(1, dtype=float)

        with torch.inference_mode():
            # Iterate over all permutations and store the saliency in sal_batch
            for result_idx, result in enumerate(trainer.forward_allperm_hr()):
                sal_batch[result_idx] = np.ndarray.item(result[2])

        # Extend the final list with the batch results
        final_sal_list.extend(sal_batch)

    return np.array(final_sal_list, dtype=float)



def save_dataset(pred_score_aesthetic, pred_score_saliency, ws_dict, batch_num, img_counter):

    score_rank_aesthetic = np.argsort(np.argsort(pred_score_aesthetic)[::-1])
    score_rank_saliency = np.argsort(np.argsort(pred_score_saliency)[::-1])
    all_score_rank = score_rank_aesthetic+score_rank_saliency

    sorted_indices = np.argsort(all_score_rank)
    result = [sorted_indices[0], sorted_indices[-1], sorted_indices[len(sorted_indices)//2]]
    
    for score_num, score_one_aesthetic in enumerate(np.sort(pred_score_aesthetic)[::-1], 1):
        ws_dict["aesthetic"].cell(img_counter, score_num).value = str(score_one_aesthetic)
    for score_num, score_one_saliency in enumerate(np.sort(pred_score_saliency)[::-1], 1):
        ws_dict["saliency"].cell(img_counter, score_num).value = str(score_one_saliency)

    for mode_num, mode in enumerate(["high", "low", "middle"]):
        img_path = opt.path_to_images+"/"+str(batch_num)+"/"+str(result[mode_num])+".jpg"
        save_path = opt.result_path+"/"+mode+"/"+str(img_counter)+".jpg"
        shutil.copyfile(img_path, save_path)

        ws_dict["result"].cell(img_counter+1, 3*mode_num+1).value = save_path
        ws_dict["result"].cell(img_counter+1, 3*mode_num+2).value = pred_score_aesthetic[result[mode_num]]
        ws_dict["result"].cell(img_counter+1, 3*mode_num+3).value = pred_score_saliency[result[mode_num]]


if __name__ =="__main__":

    #### train model
    # start_train(opt)
    #### test model
    # start_check_model(opt)
    pass